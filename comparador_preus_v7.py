"""
COMPARADOR DE PREUS SUPERMERCATS - v2 (Playwright)
===================================================
Usa Playwright (navegador real) per evitar bloquejos anti-bot.
Busca preus a Mercadona, BonPreu, Alcampo i Caprabo.

INSTAL·LACIÓ (només la primera vegada):
  pip install playwright openpyxl
  playwright install chromium

EXECUCIÓ:
  python comparador_preus_v2.py
"""

import re
import unicodedata
import os
import time
import random
from datetime import datetime
from pathlib import Path

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("Instal·lant dependències...")
    os.system("pip install playwright openpyxl")
    os.system("playwright install chromium")
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment


# ══════════════════════════════════════════════════════════════════
# CONFIGURACIÓ — edita aquí si canvien els ports
# ══════════════════════════════════════════════════════════════════

LLISTA_PATH = "llista.txt"

SUPERMERCATS_INFO = {
    "Mercadona": {"ports": 7.90,  "minim_ports_gratis": None,  "color": "00AA44"},
    "BonPreu":   {"ports": 4.90,  "minim_ports_gratis": None,  "color": "E8251F"},
    "Alcampo":   {"ports": 4.90,  "minim_ports_gratis": 49.0,  "color": "1D4289"},
    "Caprabo":   {"ports": 4.90,  "minim_ports_gratis": 60.0,  "color": "E4002B"},
}

# Posa el teu codi postal per a Mercadona (afecta disponibilitat de productes)
CODI_POSTAL_MERCADONA = "08001"


# ══════════════════════════════════════════════════════════════════
# DICCIONARI CATALÀ → CASTELLÀ
# ══════════════════════════════════════════════════════════════════

TRADUCCIONS = {
    "llet": "leche", "semidesnatada": "semidesnatada", "sencera": "entera",
    "desnatada": "desnatada", "sense": "sin", "amb": "con", "per": "para",
    "formatge": "queso", "ratllat": "rallado", "tros": "trozo",
    "macarrons": "macarrones", "tallarines": "tallarines",
    "arròs": "arroz", "farina": "harina", "sucre": "azúcar",
    "oli": "aceite", "oliva": "oliva", "gira-sol": "girasol",
    "tomàquet": "tomate", "fregit": "frito", "triturat": "triturado",
    "pèsols": "guisantes", "ceba": "cebolla", "patata": "patata",
    "nous": "nueces", "closca": "cáscara", "crues": "crudas",
    "nata": "nata", "líquida": "líquida", "cuinar": "cocinar",
    "pa": "pan", "motlle": "molde", "bastonets": "palitos",
    "higiènics": "higiénicos", "cotó": "algodón",
    "galetes": "galletas", "farcides": "rellenas", "crema": "crema",
    "xocolata": "chocolate", "cafè": "café", "gra": "grano",
    "torrat": "tostado", "te": "té", "verd": "verde",
    "menta": "menta", "poliol": "poleo", "orenga": "orégano",
    "nou": "nuez", "moscada": "moscada", "molta": "molida",
    "olives": "aceitunas", "anxova": "anchoa",
    "detergent": "detergente", "gel": "gel", "rentavaixella": "lavavajillas",
    "dosis": "dosis", "líquid": "líquido", "abrillantador": "abrillantador",
    "suavitzant": "suavizante", "paper": "papel", "forn": "horno",
    "higiènic": "higiénico", "cuina": "cocina", "ecològic": "ecológico",
    "reciclat": "reciclado", "massa": "masa", "full": "hojaldre",
    "crestes": "croissants", "base": "base", "rodona": "redonda",
    "cervesa": "cerveza", "llauna": "lata", "apta": "apta",
    "celíacs": "celíacos", "refresc": "refresco", "tònica": "tónica",
    "blau": "azul", "blanc": "blanco", "maionesa": "mayonesa",
    "puré": "puré", "patates": "patatas", "fideus": "fideos",
    "orientals": "orientales", "pollastre": "pollo", "clàssic": "clásico",
    "bossa": "bolsa", "malla": "malla", "soluble": "soluble",
    "cacau": "cacao", "gruixuts": "gruesos", "fins": "finos",
    "extra": "extra", "verge": "virgen", "refinat": "refinado",
    "suau": "suave", "blat": "trigo", "tortilles": "tortillas",
    "de": "de", "en": "en", "i": "y", "el": "el", "la": "la",
    "els": "los", "les": "las", "per": "para",
    "sense": "sin", "natural": "natural",
    "capes": "capas", "capa": "capa",
    "anxova": "anchoa", "julivert": "perejil", "all": "ajo",
    "ametlles": "almendras", "dolces": "dulces",
    "sucres": "azúcares", "afegits": "añadidos",
    "blava": "azul",
}

# Expressions multi-paraula (es substitueixen abans de la traducció paraula a paraula)
EXPRESSIONS = {
    "per a ": "para ",
    "d'oliva": "de oliva",
    "d'anxova": "de anchoa",
    "d'ametlles": "de almendras",
    "gira-sol": "girasol",
    "rentavaixelles": "lavavajillas",
    "rentavaixella": "lavavajillas",
}

MARQUES = {
    "bonpreu", "bergader", "buitoni", "colacao", "daura", "girona",
    "hellmann's", "hellmanns", "herbatural", "ismax", "schweppes",
    "sucrebo", "yatekomo", "we natural", "la collita",
    "nestlé", "nestle", "danone", "hacendado", "gallo", "barilla",
}


# ══════════════════════════════════════════════════════════════════
# UTILS
# ══════════════════════════════════════════════════════════════════

def eliminar_marca(nom: str) -> str:
    nom_lower = nom.lower().strip()
    for marca in sorted(MARQUES, key=len, reverse=True):
        if nom_lower.startswith(marca):
            nom = nom[len(marca):].strip()
            return nom.strip()
    paraules = nom.split()
    if paraules and paraules[0].isupper() and len(paraules) > 1:
        nom = " ".join(paraules[1:])
    return nom.strip()


def traduir_ca_es(text: str) -> str:
    """Tradueix català→castellà. Tracta expressions multi-paraula primer."""
    t = text.lower()
    # Expressions multi-paraula (ordre important: les més llargues primer)
    t = t.replace("per a ", "para ")
    t = t.replace("d'oliva", "de oliva")
    t = t.replace("d'anxova", "de anchoa")
    t = t.replace("d'ametlles", "de almendras")
    t = t.replace("gira-sol", "girasol")
    t = t.replace("rentavaixelles", "lavavajillas")
    t = t.replace("rentavaixella", "lavavajillas")
    # Traducció paraula a paraula
    paraules = t.split()
    resultat = []
    for p in paraules:
        p_net = re.sub(r"[^\w]", "", p)
        resultat.append(TRADUCCIONS.get(p_net, p))
    return " ".join(resultat)


def treure_accents(text: str) -> str:
    """Elimina accents: 'café' → 'cafe'"""
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


def extreure_preu(text: str) -> float | None:
    """Extreu el primer número amb decimals d'un text."""
    match = re.search(r"(\d+)[,.](\d{1,2})", text)
    if match:
        return float(f"{match.group(1)}.{match.group(2)}")
    return None


def preu_per_unitat(preu_total: float, nom_producte: str, quantitat_llista: int) -> float:
    """
    Intenta detectar si el preu trobat és d'un pack i retorna el preu unitari.
    Ex: "Leche 6x1L" preu 5.82€ → 5.82/6 = 0.97€ per unitat
    Però si la llista ja demana x2 packs, el preu ha de ser 5.82€ x2.
    Estratègia: si el nom del producte conté "NxM" o "pack de N", dividim.
    """
    # Detecta patrons com "6x1L", "pack 6", "20 unidades"
    m = re.search(r"(\d+)\s*[xX×]\s*\d", nom_producte)
    if m:
        n = int(m.group(1))
        if n > 1:
            return round(preu_total / n, 2)
    return preu_total


def pausa():
    time.sleep(random.uniform(1.5, 3.0))


# ══════════════════════════════════════════════════════════════════
# FILTRE DE RELLEVÀNCIA
# ══════════════════════════════════════════════════════════════════

def es_rellevant(nom_producte: str, terme_cerca: str, threshold: float = 0.6) -> bool:
    """
    Comprova si el nom del producte coincideix prou amb el terme cercat.
    Penalitza resultats que contenen pack NxM quan el terme no en porta.
      galetes cookies xocolata → "Galetes cookies x3"  ❌ (pack no demanat)
      galetes cookies xocolata → "Galetes cookies 200g" ✅
    """
    import re as _re
    paraules = {p for p in _re.sub(r"[^\w\s]", "", terme_cerca.lower()).split() if len(p) > 2}
    if not paraules:
        return True
    nom_lower = nom_producte.lower()

    # Penalitza: el resultat conté NxM però el terme no
    terme_te_pack = bool(_re.search(r"\d+\s*[xX]\s*\d", terme_cerca))
    resultat_te_pack = bool(_re.search(r"\d+\s*[xX]\s*\d", nom_producte))
    if resultat_te_pack and not terme_te_pack:
        return False

    coincidencies = sum(1 for p in paraules if p in nom_lower)
    return (coincidencies / len(paraules)) >= threshold


# ══════════════════════════════════════════════════════════════════
# SCRAPERS AMB PLAYWRIGHT
# ══════════════════════════════════════════════════════════════════

def cerca_mercadona(page, terme_es: str) -> float | None:
    """Cerca a Mercadona. Usa el terme en castellà."""
    try:
        url = f"https://tienda.mercadona.es/search-results?query={terme_es.replace(' ', '+')}"
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(2000)

        # Accepta cookies si apareix
        try:
            page.click("button:has-text('Aceptar')", timeout=3000)
        except Exception:
            pass

        # Busca el primer preu
        # Filtra per rellevància i agafa el mínim dels coincidents
        targetes = page.query_selector_all("[class*='product-cell'], [class*='ProductCell'], [class*='product-card']")
        preus_valids = []
        for t in targetes[:15]:
            try:
                nom_el  = t.query_selector("p[data-testid], h3, h4, [class*='name']")
                preu_el = t.query_selector('[data-testid="product-price"]')
                if nom_el and preu_el and es_rellevant(nom_el.inner_text(), terme_es):
                    p = extreure_preu(preu_el.inner_text())
                    if p and p > 0.1:
                        preus_valids.append(p)
            except Exception:
                continue
        if not preus_valids:
            el = page.query_selector('[data-testid="product-price"]')
            if el:
                p = extreure_preu(el.inner_text())
                if p and p > 0.1:
                    return p
        return min(preus_valids) if preus_valids else None
    except Exception as e:
        print(f"    ⚠️  Mercadona error: {e}")
    return None


def cerca_bonpreu(page, terme_ca: str) -> float | None:
    """Cerca a BonPreu. Usa el terme en català."""
    try:
        url = f"https://www.compraonline.bonpreuesclat.cat/search?q={terme_ca.replace(' ', '+')}"
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(2500)

        # Accepta cookies si apareix
        try:
            page.click("button:has-text('Acceptar'), button:has-text('Aceptar')", timeout=3000)
            page.wait_for_timeout(1000)
        except Exception:
            pass

        # Selectors de preu de BonPreu
        # Filtra per rellevància i agafa el mínim dels coincidents
        targetes = page.query_selector_all("[class*='sc-mmemlz']")
        preus_valids = []
        for t in targetes[:15]:
            try:
                nom_el  = t.query_selector("h3, [data-test='fop-title']")
                preu_el = t.query_selector('[data-test="fop-price"]')
                if nom_el and preu_el and es_rellevant(nom_el.inner_text(), terme_ca):
                    p = extreure_preu(preu_el.inner_text())
                    if p and p > 0.1:
                        preus_valids.append(p)
            except Exception:
                continue
        if not preus_valids:
            el = page.query_selector('[data-test="fop-price"]')
            if el:
                p = extreure_preu(el.inner_text())
                if p and p > 0.1:
                    return p
        return min(preus_valids) if preus_valids else None
    except Exception as e:
        print(f"    ⚠️  BonPreu error: {e}")
    return None


def cerca_alcampo(page, terme_es: str) -> float | None:
    """Cerca a Alcampo. URL: compraonline.alcampo.es/search?q="""
    try:
        url = f"https://www.compraonline.alcampo.es/search?q={terme_es.replace(' ', '+')}"
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(3000)

        try:
            page.click("button:has-text('Aceptar'), button:has-text('Acceptar')", timeout=3000)
            page.wait_for_timeout(1000)
        except Exception:
            pass

        try:
            page.wait_for_selector(".product-card-container", timeout=8000)
        except Exception:
            pass

        targetes = page.query_selector_all(".product-card-container")
        preus_valids = []
        for t in targetes[:15]:
            try:
                nom_el  = t.query_selector("h3, h4, [class*='name'], [class*='title']")
                preu_el = t.query_selector("[class*='price']")
                if nom_el and preu_el and es_rellevant(nom_el.inner_text(), terme_es):
                    p = extreure_preu(preu_el.inner_text())
                    if p and p > 0.1:
                        preus_valids.append(p)
            except Exception:
                continue
        if not preus_valids:
            el = page.query_selector("[class*='price']")
            if el:
                p = extreure_preu(el.inner_text())
                if p and p > 0.1:
                    return p
        return min(preus_valids) if preus_valids else None
    except Exception as e:
        print(f"    ⚠️  Alcampo error: {e}")
    return None


def cerca_caprabo(page, terme_es: str) -> float | None:
    """Cerca a Caprabo. URL: capraboacasa.com/es/search/results/?q="""
    try:
        url = f"https://www.capraboacasa.com/es/search/results/?q={terme_es.replace(' ', '+')}&suggestionsFilter=false"
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(3000)

        try:
            page.click("button:has-text('Aceptar'), button:has-text('Acceptar')", timeout=3000)
            page.wait_for_timeout(1000)
        except Exception:
            pass

        try:
            page.wait_for_selector(".product-item", timeout=8000)
        except Exception:
            pass

        targetes = page.query_selector_all(".product-item")
        preus_valids = []
        for t in targetes[:15]:
            try:
                nom_el  = t.query_selector("h3, h4, a[class*='name'], [class*='title'], [class*='product-name']")
                preu_el = t.query_selector(".price-now")
                if nom_el and preu_el and es_rellevant(nom_el.inner_text(), terme_es):
                    p = extreure_preu(preu_el.inner_text())
                    if p and p > 0.1:
                        preus_valids.append(p)
            except Exception:
                continue
        if not preus_valids:
            el = page.query_selector(".price-now")
            if el:
                p = extreure_preu(el.inner_text())
                if p and p > 0.1:
                    return p
        return min(preus_valids) if preus_valids else None
    except Exception as e:
        print(f"    ⚠️  Caprabo error: {e}")
    return None



def llegir_llista(path: str) -> list[dict]:
    productes = []
    with open(path, encoding="utf-8") as f:
        for linia in f:
            linia = linia.strip()
            if not linia or linia.startswith("#"):
                continue
            match = re.match(r"^(.+?)\s+x(\d+)\s*$", linia, re.IGNORECASE)
            if match:
                nom_original = match.group(1).strip()
                quantitat = int(match.group(2))
            else:
                nom_original = linia
                quantitat = 1

            nom_cerca_ca = treure_accents(eliminar_marca(nom_original))
            nom_cerca_es = treure_accents(traduir_ca_es(nom_cerca_ca))

            # Detecta pack NxM al nom (ex: "Llet 6x1L", "Paper higiènic 12un")
            # → el preu trobat ja és el del pack; NO multipliquem per quantitat
            te_pack = bool(re.search(r"\d+\s*[xX]\s*\d", nom_cerca_ca))

            productes.append({
                "nom": nom_original,
                "nom_cerca_ca": nom_cerca_ca,
                "nom_cerca_es": nom_cerca_es,
                "quantitat": quantitat,
                "te_pack": te_pack,
            })
    return productes


# ══════════════════════════════════════════════════════════════════
# LÒGICA PRINCIPAL
# ══════════════════════════════════════════════════════════════════

def obtenir_preus(productes: list[dict]) -> list[dict]:
    resultats = []
    total = len(productes)

    with sync_playwright() as pw:
        # Un sol navegador per tota la sessió (més eficient)
        browser = pw.chromium.launch(
            headless=True,   # Canvia a False per veure el navegador en acció
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled"]
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            locale="ca-ES",
            viewport={"width": 1280, "height": 800},
        )
        # Amaga que és Playwright (evita detecció anti-bot)
        context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => false});
            Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3]});
            window.chrome = {runtime: {}};
        """)

        # Una pàgina per supermercat (es reutilitza)
        page_mercadona  = context.new_page()
        page_bonpreu    = context.new_page()
        page_alcampo    = context.new_page()
        page_caprabo    = context.new_page()

        for i, prod in enumerate(productes, 1):
            nom          = prod["nom"]
            ca           = prod["nom_cerca_ca"]
            es           = prod["nom_cerca_es"]
            quantitat    = prod["quantitat"]
            te_pack      = prod.get("te_pack", False)
            # factor = quantitat sempre; te_pack ajuda el filtre de rellevància
            factor       = quantitat

            print(f"\n[{i}/{total}] {nom}")
            print(f"  → cerca CA: {ca}")
            print(f"  → cerca ES: {es}")

            preus = {}

            print("  Mercadona...", end=" ", flush=True)
            p = cerca_mercadona(page_mercadona, es)
            factor = 1 if te_pack else quantitat
            preus["Mercadona"] = round(p * factor, 2) if p else None
            print(f"{'%.2f€' % preus['Mercadona'] if preus['Mercadona'] else '—'}")
            pausa()

            print("  BonPreu...  ", end=" ", flush=True)
            p = cerca_bonpreu(page_bonpreu, ca)
            preus["BonPreu"] = round(p * factor, 2) if p else None
            print(f"{'%.2f€' % preus['BonPreu'] if preus['BonPreu'] else '—'}")
            pausa()

            print("  Alcampo...  ", end=" ", flush=True)
            p = cerca_alcampo(page_alcampo, es)
            preus["Alcampo"] = round(p * factor, 2) if p else None
            print(f"{'%.2f€' % preus['Alcampo'] if preus['Alcampo'] else '—'}")
            pausa()

            print("  Caprabo...  ", end=" ", flush=True)
            p = cerca_caprabo(page_caprabo, es)
            preus["Caprabo"] = round(p * factor, 2) if p else None
            print(f"{'%.2f€' % preus['Caprabo'] if preus['Caprabo'] else '—'}")
            pausa()

            preus_valids = {k: v for k, v in preus.items() if v is not None}
            mes_barat    = min(preus_valids, key=preus_valids.get) if preus_valids else "N/A"
            preu_min     = preus_valids[mes_barat] if preus_valids else None
            preu_max     = max(preus_valids.values()) if preus_valids else None

            resultats.append({
                "producte":      nom,
                "quantitat":     quantitat,
                "preus":         preus,
                "mes_barat":     mes_barat,
                "preu_mes_barat": preu_min,
                "preu_mes_car":   preu_max,
            })

        browser.close()

    return resultats


def calcular_resum(resultats: list[dict]) -> dict:
    supermercats = list(SUPERMERCATS_INFO.keys())

    total_per_super = {}
    for s in supermercats:
        total = sum(r["preus"][s] for r in resultats if r["preus"].get(s) is not None)
        total_per_super[s] = round(total, 2)

    total_fragmentat = round(
        sum(r["preu_mes_barat"] for r in resultats if r["preu_mes_barat"] is not None), 2
    )

    productes_per_super = {s: [] for s in supermercats}
    for r in resultats:
        if r["mes_barat"] in productes_per_super:
            productes_per_super[r["mes_barat"]].append(r)

    ports_per_super = {}
    for s in supermercats:
        info = SUPERMERCATS_INFO[s]
        subtotal = sum(r["preu_mes_barat"] or 0 for r in productes_per_super[s])
        if info["minim_ports_gratis"] and subtotal >= info["minim_ports_gratis"]:
            ports_per_super[s] = 0.0
        elif productes_per_super[s]:
            ports_per_super[s] = info["ports"]
        else:
            ports_per_super[s] = 0.0

    total_ports_fragmentat = sum(ports_per_super.values())
    total_real_fragmentat  = round(total_fragmentat + total_ports_fragmentat, 2)

    total_real_per_super = {}
    for s in supermercats:
        info = SUPERMERCATS_INFO[s]
        ports = 0.0 if (info["minim_ports_gratis"] and total_per_super[s] >= info["minim_ports_gratis"]) \
                    else info["ports"]
        total_real_per_super[s] = round(total_per_super[s] + ports, 2)

    super_mes_barat   = min(total_real_per_super, key=total_real_per_super.get)
    super_mes_car     = max(total_real_per_super, key=total_real_per_super.get)
    estalvi_vs_mes_car = round(total_real_per_super[super_mes_car] - total_real_fragmentat, 2)
    estalvi_fragmentant = round(total_real_per_super[super_mes_barat] - total_real_fragmentat, 2)

    return {
        "total_per_super":        total_per_super,
        "total_real_per_super":   total_real_per_super,
        "ports_per_super":        ports_per_super,
        "total_fragmentat":       total_fragmentat,
        "total_ports_fragmentat": total_ports_fragmentat,
        "total_real_fragmentat":  total_real_fragmentat,
        "productes_per_super":    productes_per_super,
        "super_mes_barat":        super_mes_barat,
        "super_mes_car":          super_mes_car,
        "estalvi_vs_mes_car":     estalvi_vs_mes_car,
        "estalvi_fragmentant":    estalvi_fragmentant,
    }


# ══════════════════════════════════════════════════════════════════
# EXPORTACIÓ EXCEL
# ══════════════════════════════════════════════════════════════════

def exportar_excel(resultats: list[dict], resum: dict, nom_fitxer: str):
    wb          = openpyxl.Workbook()
    supermercats = list(SUPERMERCATS_INFO.keys())
    colors      = {s: SUPERMERCATS_INFO[s]["color"] for s in supermercats}
    VERD        = "C6EFCE"
    GRIS        = "F2F2F2"
    GROC        = "FFEB9C"

    def cel(ws, fila, col, valor, negreta=False, fons=None, alin="center", fmt=None):
        c = ws.cell(row=fila, column=col, value=valor)
        c.font      = Font(bold=negreta, size=11)
        c.alignment = Alignment(horizontal=alin, vertical="center", wrap_text=True)
        if fons:
            c.fill = PatternFill("solid", fgColor=fons)
        if fmt:
            c.number_format = fmt
        return c

    # ── Full 1: Comparativa ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📊 Comparativa"
    ws1.row_dimensions[1].height = 38

    capceleres = ["PRODUCTE", "QUANT.", "Mercadona", "BonPreu", "Alcampo", "Caprabo", "🏆 MÉS BARAT", "💰 ESTALVI"]
    for ci, txt in enumerate(capceleres, 1):
        c = cel(ws1, 1, ci, txt, negreta=True, fons="1F3864", alin="left" if ci==1 else "center")
        c.font = Font(bold=True, color="FFFFFF", size=12)
        if ci in range(3, 7):  # Columnes supermercats
            c.fill = PatternFill("solid", fgColor=colors[supermercats[ci-3]])

    for fi, r in enumerate(resultats, 2):
        fons_fila = GRIS if fi % 2 == 0 else "FFFFFF"
        cel(ws1, fi, 1, r["producte"],  fons=fons_fila, alin="left")
        cel(ws1, fi, 2, r["quantitat"], fons=fons_fila)
        for ci, s in enumerate(supermercats, 3):
            preu = r["preus"].get(s)
            fons = VERD if s == r["mes_barat"] else fons_fila
            cel(ws1, fi, ci, preu, fons=fons, fmt='#,##0.00 "€"')
        cel(ws1, fi, 7, r["mes_barat"], fons=VERD, negreta=True)
        estalvi = round(r["preu_mes_car"] - r["preu_mes_barat"], 2) \
                  if r["preu_mes_car"] and r["preu_mes_barat"] else None
        cel(ws1, fi, 8, estalvi, fons=fons_fila, fmt='#,##0.00 "€"')

    ft = len(resultats) + 2
    cel(ws1, ft,   1, "TOTAL productes (sense ports)", negreta=True, fons=GROC, alin="left")
    for ci, s in enumerate(supermercats, 3):
        cel(ws1, ft, ci, resum["total_per_super"][s], negreta=True, fons=GROC, fmt='#,##0.00 "€"')

    fp = ft + 1
    cel(ws1, fp, 1, "PORTS", negreta=True, fons="FFC7CE", alin="left")
    for ci, s in enumerate(supermercats, 3):
        info = SUPERMERCATS_INFO[s]
        txt = f"{info['ports']}€" + (f"\n(gratis +{info['minim_ports_gratis']}€)" if info["minim_ports_gratis"] else "")
        cel(ws1, fp, ci, txt, fons="FFC7CE")

    fr = fp + 1
    ws1.row_dimensions[fr].height = 28
    cel(ws1, fr, 1, "TOTAL REAL (amb ports)", negreta=True, fons="1F3864", alin="left").font = Font(bold=True, color="FFFFFF", size=12)
    smb = resum["super_mes_barat"]
    for ci, s in enumerate(supermercats, 3):
        fons = VERD if s == smb else "FFC7CE"
        cel(ws1, fr, ci, resum["total_real_per_super"][s], negreta=True, fons=fons, fmt='#,##0.00 "€"')

    ws1.column_dimensions["A"].width = 48
    ws1.column_dimensions["B"].width = 9
    for col in ["C","D","E","F"]: ws1.column_dimensions[col].width = 15
    ws1.column_dimensions["G"].width = 15
    ws1.column_dimensions["H"].width = 14

    # ── Full 2: Per Supermercat ──────────────────────────────────
    ws2  = wb.create_sheet("🛒 Per Supermercat")
    fila = 1

    for s in supermercats:
        prods = resum["productes_per_super"][s]
        if not prods:
            continue

        ws2.row_dimensions[fila].height = 32
        c = ws2.cell(row=fila, column=1, value=f"🛒  {s.upper()}  —  compra aquí:")
        c.font      = Font(bold=True, color="FFFFFF", size=13)
        c.fill      = PatternFill("solid", fgColor=colors[s])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws2.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
        fila += 1

        for co, txt in [(1,"Producte"),(2,"Quant."),(3,"Preu total"),(4,"Preu unit.")]:
            c = ws2.cell(row=fila, column=co, value=txt)
            c.font      = Font(bold=True)
            c.fill      = PatternFill("solid", fgColor="D9D9D9")
            c.alignment = Alignment(horizontal="center", vertical="center")
        fila += 1

        subtotal = 0.0
        for r in prods:
            preu  = r["preus"].get(s) or 0
            unit  = round(preu / r["quantitat"], 2) if r["quantitat"] else preu
            subtotal += preu
            fons_fila = "F4F4F4" if fila % 2 == 0 else "FFFFFF"
            ws2.cell(row=fila, column=1, value=r["producte"]).alignment = Alignment(horizontal="left")
            ws2.cell(row=fila, column=2, value=r["quantitat"]).alignment = Alignment(horizontal="center")
            c3 = ws2.cell(row=fila, column=3, value=preu); c3.number_format = '#,##0.00 "€"'; c3.alignment = Alignment(horizontal="center")
            c4 = ws2.cell(row=fila, column=4, value=unit); c4.number_format = '#,##0.00 "€"'; c4.alignment = Alignment(horizontal="center")
            for co in range(1,5): ws2.cell(row=fila, column=co).fill = PatternFill("solid", fgColor=fons_fila)
            fila += 1

        ports = resum["ports_per_super"][s]
        for txt, val in [("Subtotal productes", round(subtotal,2)), ("Ports", ports), ("TOTAL COMANDA", round(subtotal+ports,2))]:
            ws2.cell(row=fila, column=1, value=txt).font = Font(bold=True)
            c = ws2.cell(row=fila, column=3, value=val)
            c.number_format = '#,##0.00 "€"'
            c.font = Font(bold=True)
            if txt == "TOTAL COMANDA":
                c.fill = PatternFill("solid", fgColor=VERD)
                ws2.cell(row=fila, column=1).fill = PatternFill("solid", fgColor=VERD)
            fila += 1
        fila += 2

    ws2.column_dimensions["A"].width = 48
    ws2.column_dimensions["B"].width = 9
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 15

    # ── Full 3: Resum estalvi ────────────────────────────────────
    ws3 = wb.create_sheet("💰 Resum Estalvi")

    c = ws3.cell(row=1, column=1, value="💰  RESUM D'ESTALVI MENSUAL")
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.merge_cells("A1:C1")
    ws3.row_dimensions[1].height = 42

    ws3.cell(row=2, column=1, value=f"Generat el {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(italic=True, color="808080")
    ws3.merge_cells("A2:C2")

    files = [
        ("", "", ""),
        ("OPCIÓ DE COMPRA", "COST TOTAL", "ESTALVI vs MÉS CAR"),
        (f"Tot a {resum['super_mes_car']} (el més car)", resum["total_real_per_super"][resum["super_mes_car"]], "—"),
    ]
    for s in supermercats:
        if s == resum["super_mes_car"]: continue
        estalvi_s = round(resum["total_real_per_super"][resum["super_mes_car"]] - resum["total_real_per_super"][s], 2)
        files.append((f"Tot a {s}", resum["total_real_per_super"][s], f"-{estalvi_s} €"))
    files += [
        ("", "", ""),
        (
            "🏆  FRAGMENTANT — cada producte al super més barat",
            resum["total_real_fragmentat"],
            f"-{resum['estalvi_vs_mes_car']} €  (estalvi màxim!)"
        ),
    ]

    for fi, (opcio, cost, estalvi) in enumerate(files, 3):
        es_cap  = opcio == "OPCIÓ DE COMPRA"
        es_best = "FRAGMENTANT" in str(opcio)
        fons    = "D9D9D9" if es_cap else (VERD if es_best else "FFFFFF")

        c1 = ws3.cell(row=fi, column=1, value=opcio)
        c1.font = Font(bold=es_cap or es_best, size=13 if es_best else 11)
        c1.fill = PatternFill("solid", fgColor=fons)
        c1.alignment = Alignment(horizontal="left", vertical="center")

        c2 = ws3.cell(row=fi, column=2, value=cost)
        if isinstance(cost, float): c2.number_format = '#,##0.00 "€"'
        c2.font = Font(bold=es_cap or es_best)
        c2.fill = PatternFill("solid", fgColor=fons)
        c2.alignment = Alignment(horizontal="center")

        c3 = ws3.cell(row=fi, column=3, value=estalvi)
        c3.font = Font(bold=es_cap or es_best, color="006100" if "-" in str(estalvi) else "000000")
        c3.fill = PatternFill("solid", fgColor=fons)
        c3.alignment = Alignment(horizontal="center")

    ws3.column_dimensions["A"].width = 55
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 30

    wb.save(nom_fitxer)
    print(f"\n✅ Excel guardat: {nom_fitxer}")


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("   COMPARADOR DE PREUS SUPERMERCATS  v2 (Playwright)")
    print(f"   {datetime.now().strftime('%d/%m/%Y  %H:%M')}")
    print("=" * 60)

    if not Path(LLISTA_PATH).exists():
        print(f"\n❌ No trobo '{LLISTA_PATH}'")
        print(f"   Assegura't que llista.txt és a la mateixa carpeta que aquest script.")
        input("\nPrem Enter per tancar...")
        return

    productes = llegir_llista(LLISTA_PATH)
    print(f"\n📋 {len(productes)} productes carregats\n")

    print("🌐 Iniciant navegador Playwright...")
    print("   (pot trigar 10-20 minuts per tots els productes)\n")

    resultats = obtenir_preus(productes)
    resum     = calcular_resum(resultats)

    data_str  = datetime.now().strftime("%Y%m%d")
    fitxer    = f"comparativa_preus_{data_str}.xlsx"
    exportar_excel(resultats, resum, fitxer)

    print("\n" + "=" * 60)
    print("   RESUM FINAL")
    print("=" * 60)
    for s in SUPERMERCATS_INFO:
        print(f"  {s:12} → {resum['total_real_per_super'][s]:7.2f} € (amb ports)")
    print(f"\n  🏆 FRAGMENTANT   → {resum['total_real_fragmentat']:.2f} €")
    print(f"  💰 Estalvi vs el més car:   {resum['estalvi_vs_mes_car']:.2f} €")
    print(f"  💰 Estalvi vs el més barat: {resum['estalvi_fragmentant']:.2f} €")
    print("=" * 60)

    try:
        os.startfile(fitxer)
    except Exception:
        print(f"\nObre manualment: {fitxer}")

    input("\nPrem Enter per tancar...")


if __name__ == "__main__":
    main()
